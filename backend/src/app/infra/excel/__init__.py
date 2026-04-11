"""openpyxl helpers — read and write ``.xlsx`` workbooks from/to raw bytes.

All operations are synchronous and CPU-bound. Callers that need async behavior
wrap these functions in :func:`asyncio.to_thread` (or use the higher-level
:func:`app.infra.tabular.parse_table` dispatcher which does so automatically).

Every function closes opened workbooks in ``finally`` blocks so that openpyxl's
zipfile handles never leak.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from app.core.errors import InfraError

__all__ = [
    "open_workbook",
    "read_rows",
    "write_workbook",
    "workbook_to_bytes",
]


def open_workbook(content: bytes) -> Workbook:
    """Load an openpyxl :class:`Workbook` from raw ``.xlsx`` bytes.

    Args:
        content: Raw ``.xlsx`` file content.

    Returns:
        Workbook: An openpyxl workbook opened in read/write mode.

    Raises:
        InfraError: ``SYS_002`` if the bytes are not a valid ``.xlsx``
            container.
    """
    if not isinstance(content, (bytes, bytearray)):
        raise InfraError("SYS_002", "open_workbook requires bytes content")
    buf = BytesIO(bytes(content))
    try:
        return load_workbook(buf, data_only=False)
    except (InvalidFileException, BadZipFile, KeyError, ValueError, OSError) as exc:
        raise InfraError("SYS_002", f"Invalid XLSX content: {exc}") from exc


def _resolve_sheet(workbook: Workbook, sheet_name: str | None) -> Worksheet:
    """Return the named worksheet (or the active one if ``sheet_name`` is None).

    Args:
        workbook: Source workbook.
        sheet_name: Optional worksheet name.

    Returns:
        Worksheet: The resolved worksheet.

    Raises:
        InfraError: ``SYS_002`` if the named sheet does not exist.
    """
    if sheet_name is None:
        active = workbook.active
        if active is None:
            raise InfraError("SYS_002", "Workbook has no active sheet")
        return active  # type: ignore[return-value]
    if sheet_name not in workbook.sheetnames:
        raise InfraError("SYS_002", f"Sheet not found: {sheet_name!r}")
    return workbook[sheet_name]


def read_rows(
    workbook: Workbook,
    sheet_name: str | None = None,
    *,
    header_row: int = 1,
) -> list[dict[str, Any]]:
    """Read non-empty rows from ``workbook`` as ``list[dict]``.

    Row ``header_row`` is treated as the header. Each subsequent row becomes a
    dict keyed by the header strings. Empty rows (all cells ``None`` / empty
    string) are skipped. Values are returned as-is (str/int/float/datetime/None)
    so callers can apply their own normalization.

    Args:
        workbook: Source workbook from :func:`open_workbook`.
        sheet_name: Optional worksheet name; defaults to the active sheet.
        header_row: 1-based row index of the header row. Defaults to ``1``.

    Returns:
        list[dict[str, Any]]: One dict per non-empty data row.

    Raises:
        InfraError: ``SYS_002`` if the named sheet does not exist.
    """
    if header_row < 1:
        raise InfraError("SYS_002", f"header_row must be >= 1, got {header_row}")
    sheet = _resolve_sheet(workbook, sheet_name)
    rows_iter = sheet.iter_rows(values_only=True)
    all_rows = list(rows_iter)
    if len(all_rows) < header_row:
        return []
    raw_headers = all_rows[header_row - 1]
    headers: list[str] = [
        str(cell) if cell is not None else f"column_{idx + 1}"
        for idx, cell in enumerate(raw_headers)
    ]
    output: list[dict[str, Any]] = []
    for row in all_rows[header_row:]:
        if row is None:
            continue
        if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
            continue
        record: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            record[header] = row[idx] if idx < len(row) else None
        output.append(record)
    return output


def write_workbook() -> Workbook:
    """Create a new empty :class:`Workbook` for write-out.

    Returns:
        Workbook: Fresh workbook with a single default sheet.
    """
    return Workbook()


def workbook_to_bytes(workbook: Workbook) -> bytes:
    """Serialize ``workbook`` to raw ``.xlsx`` bytes without touching disk.

    The workbook is saved into an in-memory :class:`BytesIO` buffer; the
    returned value is suitable for storing via :func:`app.infra.storage.save`
    or streaming back through an HTTP response.

    Args:
        workbook: openpyxl :class:`Workbook` to serialize.

    Returns:
        bytes: Raw ``.xlsx`` file content.

    Raises:
        InfraError: ``SYS_002`` if openpyxl raises while saving.
    """
    buf = BytesIO()
    try:
        workbook.save(buf)
    except (OSError, ValueError) as exc:
        raise InfraError("SYS_002", f"workbook_to_bytes failed: {exc}") from exc
    return buf.getvalue()
