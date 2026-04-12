"""Pure openpyxl workbook builder for M3 templates (CR-009, CR-034).

The builder is intentionally synchronous and stateless. It receives
already-filtered ``operational`` :class:`AccountCode` rows (CR-009) and
an ``actuals`` map keyed by :class:`UUID` of the account code, returns
serialized ``.xlsx`` bytes, and never touches the database or storage
layer. All DB / storage / audit side effects belong to
:class:`app.domain.templates.service.TemplateService`.

Workbook layout:

* **header** sheet — fixed-cell metadata used by the upload validator
  (FR-011 department-code check). The cell addresses match the
  spec: ``B2`` dept code, ``B3`` dept name, ``B4`` fiscal year,
  ``B5`` deadline, ``B6`` reporting currency.
* **accounts** sheet — four columns ``account_code``, ``account_name``,
  ``prior_actual``, ``budget_amount``. Rows come from
  ``operational_accounts`` in caller-supplied order. The
  ``budget_amount`` column is left blank for the filing unit to fill
  in. ``prior_actual`` is prefilled from ``actuals``; missing entries
  default to :class:`Decimal` zero (CR-034).
"""

from __future__ import annotations

from decimal import Decimal
from uuid import UUID

from openpyxl.worksheet.worksheet import Worksheet

from app.domain.accounts.models import AccountCategory, AccountCode
from app.domain.cycles.models import BudgetCycle, OrgUnit
from app.infra.excel import workbook_to_bytes, write_workbook

__all__ = ["build_template_workbook", "ACCOUNT_COLUMNS"]


#: Header row for the ``accounts`` sheet. Exposed as a module-level
#: constant so tests and the downstream upload validator (Batch 5 M4)
#: can reference the exact column labels without duplicating the list.
ACCOUNT_COLUMNS: tuple[str, ...] = (
    "account_code",
    "account_name",
    "prior_actual",
    "budget_amount",
)

_HEADER_SHEET = "header"
_ACCOUNTS_SHEET = "accounts"


def _write_header_sheet(
    sheet: Worksheet,
    *,
    cycle: BudgetCycle,
    org_unit: OrgUnit,
) -> None:
    """Populate the fixed-cell metadata sheet.

    Args:
        sheet: The openpyxl worksheet to write into.
        cycle: The source :class:`BudgetCycle` (provides fiscal year,
            deadline, currency).
        org_unit: The filing unit (provides code + name).
    """
    sheet.title = _HEADER_SHEET
    sheet["A2"] = "dept_code"
    sheet["B2"] = org_unit.code
    sheet["A3"] = "dept_name"
    sheet["B3"] = org_unit.name
    sheet["A4"] = "fiscal_year"
    sheet["B4"] = cycle.fiscal_year
    sheet["A5"] = "deadline"
    sheet["B5"] = cycle.deadline.isoformat()
    sheet["A6"] = "currency"
    sheet["B6"] = cycle.reporting_currency


def _write_accounts_sheet(
    sheet: Worksheet,
    *,
    operational_accounts: list[AccountCode],
    actuals: dict[UUID, Decimal],
) -> None:
    """Populate the account data sheet.

    Args:
        sheet: The openpyxl worksheet to populate.
        operational_accounts: Operational-category rows to write. The
            builder trusts the caller (CR-009) and does not re-filter,
            but it DOES defensively skip any non-operational row so a
            buggy caller can never leak personnel / shared_cost codes
            into the template. This is the last line of defense for the
            CR-009 invariant.
        actuals: Map of :class:`AccountCode.id` to prior actual
            :class:`Decimal`. Missing ids default to ``Decimal("0")``
            (CR-034).
    """
    sheet.title = _ACCOUNTS_SHEET
    for column_index, header in enumerate(ACCOUNT_COLUMNS, start=1):
        sheet.cell(row=1, column=column_index, value=header)

    row_index = 2
    for account in operational_accounts:
        # Reason: CR-009 defense-in-depth — builder must NEVER emit
        # personnel / shared_cost rows even if the service layer's
        # operational filter were bypassed.
        if account.category != AccountCategory.operational:
            continue
        prior_actual = actuals.get(account.id, Decimal("0"))
        sheet.cell(row=row_index, column=1, value=account.code)
        sheet.cell(row=row_index, column=2, value=account.name)
        sheet.cell(row=row_index, column=3, value=str(prior_actual))
        # Reason: budget_amount is intentionally left blank — the
        # filing unit manager fills it in before uploading (FR-011).
        sheet.cell(row=row_index, column=4, value=None)
        row_index += 1


async def build_template_workbook(
    *,
    cycle: BudgetCycle,
    org_unit: OrgUnit,
    operational_accounts: list[AccountCode],
    actuals: dict[UUID, Decimal],
) -> bytes:
    """Build a single filing unit's template workbook and return bytes.

    The function is declared ``async`` so the service layer can ``await``
    it uniformly alongside storage I/O. The actual openpyxl work is
    fully CPU-bound; for the typical ~100-account workbook it completes
    in well under a millisecond and does not require
    :func:`asyncio.to_thread` offloading.

    Args:
        cycle: Source :class:`BudgetCycle` — provides fiscal year,
            deadline, and reporting currency for the header sheet.
        org_unit: The target filing unit — provides dept code + name.
        operational_accounts: Pre-filtered list of operational-category
            :class:`AccountCode` rows in display order. Personnel and
            shared_cost rows MUST be absent (CR-009). The builder
            additionally skips any non-operational entry defensively.
        actuals: Map of :class:`AccountCode.id` to the prior-period
            :class:`Decimal` actual expense value. Missing keys map to
            :class:`Decimal` zero (CR-034).

    Returns:
        bytes: The serialized ``.xlsx`` workbook. Suitable for passing
        directly to :func:`app.infra.storage.save`.
    """
    workbook = write_workbook()
    # Reason: write_workbook() yields a default "Sheet" — rename it to
    # "header" and add the accounts sheet afterwards so the ordering is
    # deterministic (header first, accounts second).
    active = workbook.active
    if active is None:  # pragma: no cover — openpyxl always seeds one
        active = workbook.create_sheet(_HEADER_SHEET)
    _write_header_sheet(active, cycle=cycle, org_unit=org_unit)  # type: ignore[arg-type]

    accounts_sheet = workbook.create_sheet(_ACCOUNTS_SHEET)
    _write_accounts_sheet(
        accounts_sheet,
        operational_accounts=operational_accounts,
        actuals=actuals,
    )

    return workbook_to_bytes(workbook)
