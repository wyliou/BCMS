"""Unit tests for :mod:`app.domain.templates.builder` (CR-009, CR-034)."""

from __future__ import annotations

from decimal import Decimal
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.accounts.models import AccountCategory
from app.domain.templates.builder import ACCOUNT_COLUMNS, build_template_workbook
from tests.unit.templates.conftest import make_account, make_cycle, make_org_unit


def _open_bytes(content: bytes) -> tuple[Worksheet, Worksheet]:
    """Return ``(header_sheet, accounts_sheet)`` from raw ``.xlsx`` bytes."""
    from io import BytesIO

    workbook = load_workbook(BytesIO(content))
    return workbook["header"], workbook["accounts"]


async def test_build_template_contains_only_operational_codes() -> None:
    """CR-009: the builder emits one row per operational account, nothing else."""
    cycle = make_cycle()
    org_unit = make_org_unit(code="4023", name="業務部")
    accounts = [
        make_account(code="5101", name="Travel"),
        make_account(code="5102", name="Supplies"),
        make_account(code="5103", name="Utilities"),
    ]

    content = await build_template_workbook(
        cycle=cycle,
        org_unit=org_unit,
        operational_accounts=accounts,
        actuals={},
    )

    header_sheet, accounts_sheet = _open_bytes(content)
    assert header_sheet["B2"].value == "4023"
    # Header row matches the published column set.
    first_row = [
        accounts_sheet.cell(row=1, column=i + 1).value for i in range(len(ACCOUNT_COLUMNS))
    ]
    assert tuple(first_row) == ACCOUNT_COLUMNS

    # Data rows are exactly 3 (one per operational code).
    data_rows = [
        accounts_sheet.cell(row=row_index + 2, column=1).value for row_index in range(len(accounts))
    ]
    assert data_rows == ["5101", "5102", "5103"]

    # No fourth data row — builder never added phantom rows.
    assert accounts_sheet.cell(row=len(accounts) + 2, column=1).value is None


async def test_build_template_rejects_non_operational_defensively() -> None:
    """CR-009 belt-and-braces: builder skips any non-operational row it sees."""
    cycle = make_cycle()
    org_unit = make_org_unit()
    # Caller passes a mixed list by mistake — builder must defend.
    mixed = [
        make_account(code="5101"),
        make_account(code="6001", category=AccountCategory.personnel),
        make_account(code="7001", category=AccountCategory.shared_cost),
        make_account(code="5102"),
    ]

    content = await build_template_workbook(
        cycle=cycle,
        org_unit=org_unit,
        operational_accounts=mixed,
        actuals={},
    )

    _, accounts_sheet = _open_bytes(content)
    emitted_codes = []
    row_index = 2
    while True:
        code = accounts_sheet.cell(row=row_index, column=1).value
        if code is None:
            break
        emitted_codes.append(code)
        row_index += 1

    assert emitted_codes == ["5101", "5102"]


async def test_build_template_prefills_actuals_and_leaves_budget_blank() -> None:
    """Actuals dict is applied; the budget_amount column stays empty."""
    cycle = make_cycle()
    org_unit = make_org_unit()
    a1 = make_account(code="5101", name="Travel")
    a2 = make_account(code="5102", name="Supplies")
    actuals: dict[UUID, Decimal] = {a1.id: Decimal("1000.50")}

    content = await build_template_workbook(
        cycle=cycle,
        org_unit=org_unit,
        operational_accounts=[a1, a2],
        actuals=actuals,
    )

    _, accounts_sheet = _open_bytes(content)
    # Row 2 = first account
    assert accounts_sheet.cell(row=2, column=1).value == "5101"
    assert accounts_sheet.cell(row=2, column=2).value == "Travel"
    assert accounts_sheet.cell(row=2, column=3).value == "1000.50"
    assert accounts_sheet.cell(row=2, column=4).value is None  # budget blank
    # Row 3 = a2, CR-034 default zero.
    assert accounts_sheet.cell(row=3, column=3).value == "0"


async def test_build_template_zero_when_no_actual() -> None:
    """CR-034: empty actuals dict → every row's prior_actual is ``0``."""
    cycle = make_cycle()
    org_unit = make_org_unit()
    accounts = [
        make_account(code="5101"),
        make_account(code="5102"),
        make_account(code="5103"),
    ]

    content = await build_template_workbook(
        cycle=cycle,
        org_unit=org_unit,
        operational_accounts=accounts,
        actuals={},
    )

    _, accounts_sheet = _open_bytes(content)
    for row_index in range(2, 2 + len(accounts)):
        assert accounts_sheet.cell(row=row_index, column=3).value == "0"


async def test_build_template_header_populated_with_cycle_metadata() -> None:
    """Header sheet carries dept code + name + fiscal year + deadline + currency."""
    cycle = make_cycle(fiscal_year=2026, currency="TWD")
    org_unit = make_org_unit(code="4023", name="業務部")

    content = await build_template_workbook(
        cycle=cycle,
        org_unit=org_unit,
        operational_accounts=[make_account(code="5101")],
        actuals={},
    )

    header_sheet, _ = _open_bytes(content)
    assert header_sheet["B2"].value == "4023"
    assert header_sheet["B3"].value == "業務部"
    assert header_sheet["B4"].value == 2026
    assert header_sheet["B5"].value == cycle.deadline.isoformat()
    assert header_sheet["B6"].value == "TWD"


async def test_build_template_zero_operational_accounts_still_builds() -> None:
    """Zero operational accounts → header + empty accounts sheet, no error."""
    cycle = make_cycle()
    org_unit = make_org_unit()

    content = await build_template_workbook(
        cycle=cycle,
        org_unit=org_unit,
        operational_accounts=[],
        actuals={},
    )

    header_sheet, accounts_sheet = _open_bytes(content)
    assert header_sheet["B2"].value == org_unit.code
    # Row 1 = header; row 2 must be empty.
    assert accounts_sheet.cell(row=2, column=1).value is None
    # Header row still written.
    assert accounts_sheet.cell(row=1, column=1).value == "account_code"
